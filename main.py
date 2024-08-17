"""
Coffee Shop Analysis
"""


import datetime
import os
import collections
import json

import openpyxl


Order = collections.namedtuple(
    'Order', 
    [
        'order_id', 'order_date', 'customer_id', 
        'product_id', 'quantity'
    ]
)

Customer = collections.namedtuple(
    'Customer',
    [
        'customer_id', 'customer_name', 'email',
        'phone_number', 'address', 'city', 'country',
        'post_code', 'loyalty_card'
    ]
)

Product = collections.namedtuple(
    'Product',
    [
        'product_id', 'coffee_type', 'roast_type', 'size', 
        'unit_price', 'price_per_100g', 'profit',
    ]
)


ROAST_TYPE = {
    'L': 'Light',
    'M': 'Medium',
    'D': 'Dark',
}

COFFEE_TYPE = {
    'Ara': 'Arabica',
    'Rob': 'Robusta',
    'Lib': 'Liberica', 
    'Exc': 'Excelsa',
}

def read_data(workbook_path):
    """Read data from Excel Workbook

    Excel Workbook should contain three worksheets of data:
        orders
        customers
        products
    """
    wb = openpyxl.load_workbook(workbook_path)
    orders_sheet = wb['orders']
    order_data = []
    for row in orders_sheet.iter_rows():
        # First row is headers, skip it
        if row[0].value == 'Order ID':
            continue

        order_id = row[0].value
        order_date_text = row[1].value
        customer_id = row[2].value
        product_id = row[3].value
        quantity = int(row[4].value) if row[4].value else 0
        order_date = order_date_text
        order = Order(order_id, order_date, customer_id, product_id, quantity)
        order_data.append(order)

    customer_data = dict()
    customer_sheet = wb['customers']
    for row in customer_sheet.iter_rows():
        if row[0].value == 'Customer ID':
            continue
        customer_id = row[0].value
        customer_name = row[1].value
        email = row[2].value
        phone_number = row[3].value
        address = row[4].value 
        city = row[5].value 
        country = row[6].value
        post_code = row[7].value
        loyalty_card = row[8].value
        customer = Customer(
            customer_id, customer_name, email, phone_number,
            address, city, country, post_code, loyalty_card
        )
        customer_data[customer.customer_id] = customer

    product_data = dict()
    product_sheet = wb['products']
    for row in product_sheet.iter_rows():
        if row[0].value == 'Product ID':
            continue
        
        product_id = row[0].value
        coffee_type_code = row[1].value
        roast_type_code = row[2].value
        size = row[3].value
        unit_price = row[4].value
        price_per_100g = row[5].value
        profit = row[6].value
        product = Product(
            product_id, COFFEE_TYPE[coffee_type_code],
            ROAST_TYPE[roast_type_code], size,
            unit_price, price_per_100g, profit
        )
        product_data[product.product_id] = product

    return order_data, customer_data, product_data

    wb.close()


def join_data(order_data, customer_data, product_data):
    """Joins the data into a single dataset.
    """
    joined_data = []
    for order in order_data:
        customer = customer_data[order.customer_id]
        product = product_data[order.product_id]
        record = [
            datetime.datetime.strftime(order.order_date, '%Y-%m'), 
            customer.customer_name,
            customer.country,
            customer.loyalty_card,
            product.coffee_type,
            product.roast_type,
            product.size,
            round(order.quantity * product.unit_price, 2)
        ]
        joined_data.append(record)
    return joined_data


def main():
    "Main program"
    workbook_path = os.path.join('.', 'data', 'coffeeOrdersData.xlsx')
    order_data, customer_data, product_data = read_data(workbook_path)
    output_data = join_data(order_data, customer_data, product_data)
    with open(os.path.join('.', 'output', 'data.js'), 'w') as outfile:
        outfile.write('// Data file for coffee analysis\n')
        coffee_types = list(COFFEE_TYPE.values())
        outfile.write(f'const coffeeTypes = {json.dumps(coffee_types)};\n')
        date_series = list(sorted(set(t[0] for t in output_data)))
        outfile.write(f'const dateSeries = {json.dumps(date_series)};\n')
        outfile.write(f'const coffeeData = {json.dumps(output_data)};\n')


if __name__ == '__main__':
    main()